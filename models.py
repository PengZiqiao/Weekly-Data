from openpyxl import load_workbook

FILE = 'E:/weekly/static/供销价.xlsx'
SHEET_LIST = ['商业', '办公', '住宅', '别墅']


def open_sheet(sheet):
    """打开对应sheet，并获取 '合计' 所在行编号"""
    wb = load_workbook(FILE)
    ws = wb[sheet]
    i = 1
    while not ws["a" + str(i)].value == '合计':
        i += 1
    return ws, str(i)


def data_to_dict(sheet):
    """将数据储存在字典中， # data的结构为 [键1:[值, 环比],键2:[值, 环比],...]"""
    ws, i = open_sheet(sheet)
    data = dict()
    mapping = [
        # 键，   值， 环比
        ['上市', 'b', 'c'],
        ['认购', 'f', 'g'],
        ['成交', 'j', 'k'],
    ]
    for each in mapping:
        # 先添加面积类的数据
        key_name = each[0]
        value_col = each[1]
        ratio_col = each[2]
        data[key_name] = list()  # 申明一个列表，结构为 [值,环比]
        value = float(ws[value_col + i].value) / 10000  # 面积都换算成“万方”为单位
        data[key_name].append(format(value, '.2f'))
        ratio = ws[ratio_col + i].value
        if ratio == 'Na':
            data[key_name].append('(无环比)')
        else:
            data[key_name].append(['上涨', '下滑'][float(ratio) < 0] + str(abs(ratio)))
    # 再添加价格数据
    data['均价'] = list()
    data['均价'].append(ws['p' + i].value)  # 值
    ratio = ws['q' + i].value  # 环比
    if ratio == 'Na':
        data[key_name].append('(无环比)')
    else:
        data['均价'].append(['上涨', '下滑'][float(ratio) < 0] + str(abs(ratio)))
    return data


def get_data():
    """主程序部分，返回一个完整的嵌套字典，结构为{ 住宅:{ 上市:[ 值, 环比]...}}"""
    data = dict()
    for sheet in SHEET_LIST:
        data[sheet] = data_to_dict(sheet)
    return data
